#!/usr/bin/env python3
"""
Скрипт тестирования экспорта УПД в Excel
Проверяет все импорты, логику и генерацию файлов
"""

import sys
import json
from pathlib import Path
import traceback

# Добавляем путь к приложению
sys.path.insert(0, str(Path(__file__).parent))

print("=" * 80)
print("ТЕСТИРОВАНИЕ ЭКСПОРТА УПД В EXCEL")
print("=" * 80)

# Шаг 1: Проверка импортов
print("\n[1/6] Проверка импортов...")
try:
    from app.services.excel_export import ExcelExportService
    print("✅ ExcelExportService импортирован")
    
    import openpyxl
    print(f"✅ openpyxl версия {openpyxl.__version__}")
    
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    print("✅ openpyxl.styles импортированы")
    
    from jinja2 import Environment, FileSystemLoader
    print("✅ Jinja2 импортирован")
    
except Exception as e:
    print(f"❌ Ошибка импорта: {e}")
    traceback.print_exc()
    sys.exit(1)

# Шаг 2: Поиск тестовых документов
print("\n[2/6] Поиск тестовых документов...")
DOCUMENTS_DIR = Path(__file__).parent / "documents"

upd_docs = []
for doc_folder in DOCUMENTS_DIR.iterdir():
    if doc_folder.is_dir():
        metadata_path = doc_folder / "metadata.json"
        if metadata_path.exists():
            try:
                metadata = json.loads(metadata_path.read_text(encoding='utf-8'))
                if metadata.get('type') == 'upd':
                    upd_docs.append({
                        'id': metadata['id'],
                        'user_id': metadata.get('user_id'),
                        'number': metadata.get('document_number'),
                        'date': metadata.get('document_date'),
                        'seller': metadata.get('seller_name'),
                        'buyer': metadata.get('buyer_name'),
                        'total': metadata.get('total_amount')
                    })
            except Exception as e:
                continue

if upd_docs:
    print(f"✅ Найдено {len(upd_docs)} УПД документов")
    test_doc = upd_docs[0]
    print(f"   Используем для теста: {test_doc['id']}")
    print(f"   № {test_doc['number']} от {test_doc['date']}")
    print(f"   {test_doc['seller']} → {test_doc['buyer']}")
    print(f"   Сумма: {test_doc['total']} руб.")
else:
    print("❌ УПД документы не найдены")
    sys.exit(1)

# Шаг 3: Проверка структуры документа
print("\n[3/6] Проверка структуры документа...")
doc_folder = DOCUMENTS_DIR / test_doc['id']
required_files = ['metadata.json', 'document.html']
optional_files = ['form_data.json', 'data.json']

for file in required_files:
    file_path = doc_folder / file
    if file_path.exists():
        size = file_path.stat().st_size
        print(f"✅ {file} ({size} bytes)")
    else:
        print(f"❌ {file} отсутствует")
        sys.exit(1)

for file in optional_files:
    file_path = doc_folder / file
    if file_path.exists():
        size = file_path.stat().st_size
        print(f"✅ {file} ({size} bytes)")

# Шаг 4: Загрузка данных документа
print("\n[4/6] Загрузка данных документа...")
try:
    # Читаем form_data
    form_data_path = doc_folder / "form_data.json"
    if not form_data_path.exists():
        form_data_path = doc_folder / "data.json"
    
    if form_data_path.exists():
        form_data = json.loads(form_data_path.read_text(encoding='utf-8'))
        print(f"✅ Данные формы загружены")
        print(f"   Товаров/услуг: {len(form_data.get('items', []))}")
        print(f"   Итого без НДС: {form_data.get('total_amount_without_vat', 0)}")
        print(f"   НДС: {form_data.get('total_vat_amount', 0)}")
        print(f"   Итого с НДС: {form_data.get('total_amount_with_vat', 0)}")
    else:
        print("⚠️  form_data.json не найден, будет использован HTML")
        form_data = {}
    
    # Читаем HTML
    html_path = doc_folder / "document.html"
    html_content = html_path.read_text(encoding='utf-8')
    print(f"✅ HTML документ загружен ({len(html_content)} символов)")
    
except Exception as e:
    print(f"❌ Ошибка загрузки данных: {e}")
    traceback.print_exc()
    sys.exit(1)

# Шаг 5: Тестирование генерации XLS
print("\n[5/6] Тестирование генерации XLS...")
try:
    service = ExcelExportService()
    
    # Тестируем через приватный метод (для отладки)
    metadata = json.loads((doc_folder / "metadata.json").read_text(encoding='utf-8'))
    
    # Генерируем XLS
    enhanced_html = service._enhance_html_for_excel(html_content)
    print(f"✅ HTML подготовлен для Excel ({len(enhanced_html)} символов)")
    
    filename = service._get_excel_filename(metadata, 'xls')
    print(f"✅ Имя файла: {filename}")
    
    # Проверяем наличие Office XML тегов
    if 'xmlns:x="urn:schemas-microsoft-com:office:excel"' in html_content or '<table' in html_content:
        print("✅ HTML содержит таблицы для Excel")
    else:
        print("⚠️  HTML может не содержать подходящих тегов для Excel")
    
except Exception as e:
    print(f"❌ Ошибка генерации XLS: {e}")
    traceback.print_exc()

# Шаг 6: Тестирование генерации XLSX
print("\n[6/6] Тестирование генерации XLSX...")
try:
    if form_data:
        # Генерируем XLSX
        buffer = service._create_xlsx_from_upd_data(form_data)
        buffer_size = buffer.tell()
        buffer.seek(0)
        print(f"✅ XLSX файл создан ({buffer_size} bytes)")
        
        # Сохраняем для проверки
        test_file = Path("/tmp/test_upd.xlsx")
        test_file.write_bytes(buffer.read())
        print(f"✅ Тестовый файл сохранен: {test_file}")
        
        # Проверяем что файл валидный
        from openpyxl import load_workbook
        wb = load_workbook(test_file)
        ws = wb.active
        print(f"✅ Файл валидный, листов: {len(wb.worksheets)}")
        print(f"   Активный лист: '{ws.title}'")
        print(f"   Размер: {ws.max_row} строк x {ws.max_column} колонок")
        
        # Проверяем несколько ключевых ячеек
        if ws['A1'].value:
            print(f"   Заголовок A1: {ws['A1'].value[:50]}...")
        
        print("\n📋 Структура документа:")
        for row in range(1, min(15, ws.max_row + 1)):
            cell_a = ws.cell(row, 1).value
            if cell_a:
                print(f"   Строка {row}: {str(cell_a)[:60]}")
        
    else:
        print("⚠️  Нет данных формы для генерации XLSX")
        
except Exception as e:
    print(f"❌ Ошибка генерации XLSX: {e}")
    traceback.print_exc()
    import traceback
    traceback.print_exc()

# Итоги
print("\n" + "=" * 80)
print("РЕЗУЛЬТАТЫ ТЕСТИРОВАНИЯ")
print("=" * 80)
print(f"""
Документ ID: {test_doc['id']}
User ID: {test_doc['user_id']}
Номер: {test_doc['number']}
Дата: {test_doc['date']}

Для тестирования через API используйте:

1. XLS экспорт:
curl -X GET "http://localhost:8000/api/v1/documents/saved/{test_doc['id']}/export-excel?format=xls" \\
  -H "Authorization: Bearer YOUR_TOKEN" \\
  --output test_upd.xls

2. XLSX экспорт:
curl -X GET "http://localhost:8000/api/v1/documents/saved/{test_doc['id']}/export-excel?format=xlsx" \\
  -H "Authorization: Bearer YOUR_TOKEN" \\
  --output test_upd.xlsx

Или используйте через браузер (с активной сессией):
http://localhost:8000/api/v1/documents/saved/{test_doc['id']}/export-excel?format=xlsx
""")
print("=" * 80)
