#!/usr/bin/env python3
"""
Тестирование API endpoint экспорта в Excel
"""

import requests
import json
import sys
from pathlib import Path

BASE_URL = "http://localhost:8000"
DOCUMENT_ID = "42f4cbc3-1ab4-4586-80e1-73704feb4996"
USER_ID = 6

print("=" * 80)
print("ТЕСТИРОВАНИЕ API ЭКСПОРТА В EXCEL")
print("=" * 80)

# Шаг 1: Попытка получить токен (тестовый вход)
print("\n[1/5] Получение токена авторизации...")

# Проверяем существующих пользователей
try:
    # Попробуем войти с тестовыми данными
    # Обычно тестовый пользователь: test@documatica.ru / test123
    login_data = {
        "username": "test@documatica.ru",
        "password": "test123"
    }
    
    response = requests.post(f"{BASE_URL}/api/v1/auth/login", data=login_data)
    
    if response.status_code == 200:
        token = response.json().get("access_token")
        print(f"✅ Токен получен: {token[:20]}...")
    else:
        print(f"⚠️  Не удалось войти с тестовыми данными")
        print(f"   Код: {response.status_code}")
        
        # Попробуем без авторизации (должна быть ошибка 401)
        print("\n   Попробуем запрос без авторизации (ожидаем 401)...")
        token = None
        
except Exception as e:
    print(f"⚠️  Ошибка получения токена: {e}")
    print("   Продолжаем тестирование без токена...")
    token = None

# Шаг 2: Тест без авторизации (должен вернуть 401)
print("\n[2/5] Тест без авторизации (ожидаем 401)...")
try:
    response = requests.get(
        f"{BASE_URL}/api/v1/documents/saved/{DOCUMENT_ID}/export-excel",
        params={"format": "xlsx"}
    )
    
    if response.status_code == 401:
        print(f"✅ Корректно вернул 401 Unauthorized")
    else:
        print(f"⚠️  Неожиданный код: {response.status_code}")
        print(f"   Ответ: {response.text[:200]}")
        
except Exception as e:
    print(f"❌ Ошибка запроса: {e}")

# Шаг 3: Тест с несуществующим документом (ожидаем 404)
print("\n[3/5] Тест с несуществующим документом (ожидаем 404)...")
if token:
    try:
        response = requests.get(
            f"{BASE_URL}/api/v1/documents/saved/00000000-0000-0000-0000-000000000000/export-excel",
            params={"format": "xlsx"},
            headers={"Authorization": f"Bearer {token}"}
        )
        
        if response.status_code == 404:
            print(f"✅ Корректно вернул 404 Not Found")
        else:
            print(f"⚠️  Неожиданный код: {response.status_code}")
            
    except Exception as e:
        print(f"❌ Ошибка запроса: {e}")
else:
    print("⚠️  Пропущен (нет токена)")

# Шаг 4: Тест с неверным форматом (ожидаем 400)
print("\n[4/5] Тест с неверным форматом (ожидаем 400)...")
if token:
    try:
        response = requests.get(
            f"{BASE_URL}/api/v1/documents/saved/{DOCUMENT_ID}/export-excel",
            params={"format": "pdf"},
            headers={"Authorization": f"Bearer {token}"}
        )
        
        if response.status_code == 400:
            print(f"✅ Корректно вернул 400 Bad Request")
        else:
            print(f"⚠️  Код: {response.status_code}")
            print(f"   Ответ: {response.text[:200]}")
            
    except Exception as e:
        print(f"❌ Ошибка запроса: {e}")
else:
    print("⚠️  Пропущен (нет токена)")

# Шаг 5: Реальное тестирование экспорта
print("\n[5/5] Тестирование экспорта файлов...")

if token:
    # Тест XLS
    print("\n   📄 Экспорт в XLS...")
    try:
        response = requests.get(
            f"{BASE_URL}/api/v1/documents/saved/{DOCUMENT_ID}/export-excel",
            params={"format": "xls"},
            headers={"Authorization": f"Bearer {token}"}
        )
        
        if response.status_code == 200:
            content_type = response.headers.get('Content-Type')
            content_length = len(response.content)
            
            print(f"   ✅ Статус: {response.status_code}")
            print(f"   ✅ Content-Type: {content_type}")
            print(f"   ✅ Размер: {content_length} bytes")
            
            # Сохраняем файл
            output_file = Path("/tmp/test_api_upd.xls")
            output_file.write_bytes(response.content)
            print(f"   ✅ Сохранен: {output_file}")
            
            # Проверяем что это HTML (для XLS)
            if b'<table' in response.content or b'<html' in response.content:
                print(f"   ✅ Содержимое: HTML для Excel")
            else:
                print(f"   ⚠️  Неожиданный формат содержимого")
        else:
            print(f"   ❌ Ошибка: {response.status_code}")
            print(f"   {response.text[:300]}")
            
    except Exception as e:
        print(f"   ❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
    
    # Тест XLSX
    print("\n   📊 Экспорт в XLSX...")
    try:
        response = requests.get(
            f"{BASE_URL}/api/v1/documents/saved/{DOCUMENT_ID}/export-excel",
            params={"format": "xlsx"},
            headers={"Authorization": f"Bearer {token}"}
        )
        
        if response.status_code == 200:
            content_type = response.headers.get('Content-Type')
            content_length = len(response.content)
            
            print(f"   ✅ Статус: {response.status_code}")
            print(f"   ✅ Content-Type: {content_type}")
            print(f"   ✅ Размер: {content_length} bytes")
            
            # Сохраняем файл
            output_file = Path("/tmp/test_api_upd.xlsx")
            output_file.write_bytes(response.content)
            print(f"   ✅ Сохранен: {output_file}")
            
            # Проверяем что это действительно XLSX (magic bytes)
            if response.content[:4] == b'PK\x03\x04':
                print(f"   ✅ Формат: Валидный ZIP/XLSX (magic bytes)")
                
                # Пробуем открыть через openpyxl
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(output_file)
                    ws = wb.active
                    print(f"   ✅ Файл открывается в openpyxl")
                    print(f"   ✅ Листов: {len(wb.worksheets)}, активный: '{ws.title}'")
                    print(f"   ✅ Размер: {ws.max_row} строк x {ws.max_column} колонок")
                except Exception as e:
                    print(f"   ⚠️  Не удалось открыть через openpyxl: {e}")
            else:
                print(f"   ⚠️  Неожиданный формат (первые 4 байта: {response.content[:4]})")
        else:
            print(f"   ❌ Ошибка: {response.status_code}")
            print(f"   {response.text[:300]}")
            
    except Exception as e:
        print(f"   ❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        
else:
    print("⚠️  Невозможно протестировать экспорт без токена")
    print("\nПопробуйте вручную через curl с вашим токеном:")
    print(f'\ncurl -X GET "http://localhost:8000/api/v1/documents/saved/{DOCUMENT_ID}/export-excel?format=xlsx" \\')
    print(f'  -H "Authorization: Bearer YOUR_TOKEN" \\')
    print(f'  --output test_upd.xlsx')

print("\n" + "=" * 80)
print("ИТОГИ ТЕСТИРОВАНИЯ")
print("=" * 80)
