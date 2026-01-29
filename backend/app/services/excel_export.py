"""
Сервис экспорта документов в Excel (XLS/XLSX)

Реализует два подхода:
- XLS: Excel 2003 XML (SpreadsheetML) с полной поддержкой Unicode и объединения ячеек
- XLSX: openpyxl (качественный формат)
"""

import io
import json
import logging
import html
from pathlib import Path
from typing import Dict, Optional, BinaryIO, Tuple
from datetime import datetime

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
import xlwt  # Оставляем для akt/invoice пока
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from jinja2 import Environment, FileSystemLoader

# Настройка логирования
logger = logging.getLogger(__name__)

# Пути к файлам
TEMPLATES_DIR = Path(__file__).parent.parent / "templates"
DOCUMENTS_DIR = Path(__file__).parent.parent.parent / "documents"

# Настройка Jinja2 для генерации HTML
jinja_env = Environment(
    loader=FileSystemLoader(TEMPLATES_DIR),
    autoescape=True
)


class ExcelExportService:
    """Сервис конвертации документов в Excel"""
    
    SUPPORTED_TYPES = ['upd', 'akt', 'invoice']
    
    def __init__(self):
        """Инициализация сервиса"""
        self.documents_dir = DOCUMENTS_DIR
        self.templates_dir = TEMPLATES_DIR
    
    def export_to_xls(self, document_id: str, user_id: int) -> StreamingResponse:
        """
        Экспорт документа в XLS через XML-шаблон (Excel 2003 SpreadsheetML)
        с методом write_merge() для создания реальных объединенных ячеек.
        
        Args:
            document_id: UUID документа
            user_id: ID пользователя (для проверки прав)
        
        Returns:
            StreamingResponse с XLS файлом
        
        Raises:
            HTTPException: 403, 404, 422, 500
        """
        logger.info(f"XLS export requested: document_id={document_id}, user_id={user_id}")
        
        try:
            # Читаем данные документа и проверяем права
            metadata, form_data, _ = self._read_document_data(document_id, user_id)
            
            # Проверяем тип документа
            doc_type = metadata.get('type')
            if doc_type not in self.SUPPORTED_TYPES:
                raise HTTPException(
                    status_code=422,
                    detail=f"Экспорт в Excel для типа '{doc_type}' не поддерживается"
                )
            
            # Создаем XLS в зависимости от типа документа
            if doc_type == 'upd':
                buffer = self._create_xls_from_upd_data(form_data)
            elif doc_type == 'akt':
                buffer = self._create_xls_from_akt_data(form_data)
            elif doc_type == 'invoice':
                buffer = self._create_xls_from_invoice_data(form_data)
            else:
                raise HTTPException(
                    status_code=422,
                    detail=f"XLS экспорт для типа '{doc_type}' еще не реализован"
                )
            
            # Формируем имя файла
            filename = self._get_excel_filename(metadata, 'xls')
            
            logger.info(f"XLS export completed: document_id={document_id}, filename={filename}")
            
            return StreamingResponse(
                buffer,
                media_type="application/vnd.ms-excel",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Cache-Control": "no-cache, no-store, must-revalidate",
                    "Pragma": "no-cache",
                    "Expires": "0"
                }
            )
            
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"XLS export error: {str(e)}, document_id={document_id}")
            raise HTTPException(
                status_code=500,
                detail=f"Ошибка экспорта в XLS: {str(e)}"
            )
    
    def export_to_xlsx(self, document_id: str, user_id: int) -> StreamingResponse:
        """
        Экспорт документа в XLSX через openpyxl
        
        Этап 2: Качественный подход с точным контролем форматирования.
        Создает настоящий XLSX файл с правильной структурой, стилями и формулами.
        
        Args:
            document_id: UUID документа
            user_id: ID пользователя (для проверки прав)
        
        Returns:
            StreamingResponse с XLSX файлом
        
        Raises:
            HTTPException: 403, 404, 422, 500
        """
        logger.info(f"XLSX export requested: document_id={document_id}, user_id={user_id}")
        
        try:
            # Читаем данные документа и проверяем права
            metadata, form_data, _ = self._read_document_data(document_id, user_id)
            
            # Проверяем тип документа
            doc_type = metadata.get('type')
            if doc_type not in self.SUPPORTED_TYPES:
                raise HTTPException(
                    status_code=422,
                    detail=f"Экспорт в Excel для типа '{doc_type}' не поддерживается"
                )
            
            # Создаем XLSX в зависимости от типа документа
            if doc_type == 'upd':
                buffer = self._create_xlsx_from_upd_data(form_data)
            elif doc_type == 'akt':
                buffer = self._create_xlsx_from_akt_data(form_data)
            elif doc_type == 'invoice':
                buffer = self._create_xlsx_from_invoice_data(form_data)
            else:
                raise HTTPException(
                    status_code=422,
                    detail=f"XLSX экспорт для типа '{doc_type}' еще не реализован"
                )
            
            # Формируем имя файла
            filename = self._get_excel_filename(metadata, 'xlsx')
            
            logger.info(f"XLSX export completed: document_id={document_id}, filename={filename}")
            
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )
            
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"XLSX export error: {str(e)}, document_id={document_id}")
            raise HTTPException(
                status_code=500,
                detail=f"Ошибка экспорта в XLSX: {str(e)}"
            )
    
    def _read_document_data(
        self, 
        document_id: str, 
        user_id: int
    ) -> Tuple[Dict, Dict, Optional[str]]:
        """
        Чтение данных документа с проверкой прав доступа
        
        Args:
            document_id: UUID документа
            user_id: ID пользователя
        
        Returns:
            Tuple[metadata, form_data, html_content]
        
        Raises:
            HTTPException: 403, 404
        """
        doc_folder = self.documents_dir / document_id
        
        # Проверка существования документа
        if not doc_folder.exists():
            raise HTTPException(status_code=404, detail="Документ не найден")
        
        # Загрузка метаданных
        metadata_path = doc_folder / "metadata.json"
        if not metadata_path.exists():
            raise HTTPException(status_code=404, detail="Метаданные документа не найдены")
        
        metadata = json.loads(metadata_path.read_text(encoding='utf-8'))
        
        # Проверка владельца
        if metadata.get('user_id') != user_id:
            logger.warning(
                f"Access denied: user {user_id} tried to access document {document_id} "
                f"(owner: {metadata.get('user_id')})"
            )
            raise HTTPException(status_code=403, detail="Доступ запрещен")
        
        # Загрузка form_data
        form_data_path = doc_folder / "form_data.json"
        if not form_data_path.exists():
            # Для старых документов может не быть form_data, пробуем data.json
            form_data_path = doc_folder / "data.json"
        
        if form_data_path.exists():
            form_data = json.loads(form_data_path.read_text(encoding='utf-8'))
        else:
            form_data = {}
        
        # Загрузка HTML (опционально)
        html_path = doc_folder / "document.html"
        html_content = None
        if html_path.exists():
            html_content = html_path.read_text(encoding='utf-8')
        
        return metadata, form_data, html_content
    
    def _generate_html_from_data(self, doc_type: str, form_data: Dict) -> str:
        """
        Генерация HTML из form_data (если HTML файл отсутствует)
        
        Args:
            doc_type: Тип документа (upd, akt, invoice)
            form_data: Данные формы
        
        Returns:
            HTML строка
        """
        # Выбор шаблона
        template_map = {
            'upd': 'upd_template.html',
            'akt': 'akt_template.html',
            'invoice': 'invoice_template.html'
        }
        
        template_name = template_map.get(doc_type)
        if not template_name:
            raise ValueError(f"Неизвестный тип документа: {doc_type}")
        
        template = jinja_env.get_template(template_name)
        
        # Подготовка данных для рендеринга (упрощенная версия)
        # В реальности здесь нужна полная логика как в documents.py
        template_data = form_data.copy()
        
        # Конвертируем Decimal в float для Jinja2
        if 'items' in template_data:
            for item in template_data['items']:
                for key in ['quantity', 'price', 'amount_without_vat', 'vat_amount', 'amount_with_vat', 'amount']:
                    if key in item and item[key] is not None:
                        item[key] = float(item[key])
        
        html_content = template.render(**template_data)
        return html_content
    
    def _enhance_html_for_excel(self, html_content: str) -> str:
        """
        Улучшение HTML для лучшей совместимости с Excel
        
        Добавляет/корректирует теги для оптимального отображения в Excel.
        В текущей реализации шаблоны уже оптимизированы, поэтому просто возвращаем как есть.
        
        Args:
            html_content: Исходный HTML
        
        Returns:
            Улучшенный HTML
        """
        # Шаблоны уже содержат xmlns:x="urn:schemas-microsoft-com:office:excel"
        # и другие необходимые теги, поэтому просто возвращаем как есть
        return html_content
    
    def _create_xls_from_upd_data(self, form_data: Dict) -> BinaryIO:
        """
        Создание XLS файла для УПД через XML-шаблон (Excel 2003 XML/SpreadsheetML)
        
        Использует upd_excel_template.xml с Jinja2 переменными для генерации
        идеального УПД формата с полным сохранением всех стилей, объединений
        и структуры из эталонного файла upd_11_clean.xml.
        
        Args:
            form_data: Данные формы УПД
        
        Returns:
            BytesIO с XML содержимым (Excel откроет его как .xls)
        """
        logger.info("🔥 CREATING XLS VIA XML TEMPLATE (Excel 2003 SpreadsheetML)")
        
        try:
            # Загружаем Jinja2 шаблон
            template = jinja_env.get_template('upd_excel_template.xml')
            
            # Подготавливаем данные для рендеринга
            template_data = {
                'document_number': form_data.get('document_number', ''),
                'document_date': form_data.get('document_date', ''),
                'correction_number': form_data.get('correction_number') or '',
                'correction_date': form_data.get('correction_date') or '',
                'status': form_data.get('status', 1),
                'seller': form_data.get('seller', {}),
                'buyer': form_data.get('buyer', {}),
                'consignor': form_data.get('consignor') or form_data.get('seller', {}),
                'consignee': form_data.get('consignee') or form_data.get('buyer', {}),
                'items': form_data.get('items', []),
                'total_amount_without_vat': form_data.get('total_amount_without_vat', '0'),
                'total_vat_amount': form_data.get('total_vat_amount', '0'),
                'total_amount_with_vat': form_data.get('total_amount_with_vat', '0'),
                'currency_name': form_data.get('currency_name', 'Российский рубль'),
                'currency_code': form_data.get('currency_code', '643'),
                'gov_contract_id': form_data.get('gov_contract_id') or '',
                'payment_document': form_data.get('payment_document') or '',
                'shipping_document': form_data.get('shipping_document') or '',
                'seller_signer': form_data.get('seller_signer', {}),
            }
            
            # Рендерим XML через Jinja2
            xml_content = template.render(**template_data)
            
            # Создаем BytesIO буфер с результатом
            buffer = io.BytesIO()
            buffer.write(xml_content.encode('utf-8'))
            buffer.seek(0)
            
            logger.info("✅ XLS (XML) generated successfully via template")
            return buffer
            
        except Exception as e:
            logger.error(f"❌ Error generating XLS from template: {str(e)}")
            raise
        
        
    
    
    def _create_xls_from_akt_data(self, form_data: Dict) -> BinaryIO:
        """
        Создание XLS файла для Акта выполненных работ через xlwt
        
        Args:
            form_data: Данные формы Акта
        
        Returns:
            BytesIO с XLS файлом
        """
        # Упрощенная реализация для Акта (можно расширить позже)
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Акт')
        
        style_title = xlwt.XFStyle()
        font_title = xlwt.Font()
        font_title.name = 'Arial'
        font_title.bold = True
        font_title.height = 280  # 14pt
        style_title.font = font_title
        
        ws.write(0, 0, 'АКТ ВЫПОЛНЕННЫХ РАБОТ (УСЛУГ)', style_title)
        ws.write(2, 0, f"№ {form_data.get('document_number', '')}")
        ws.write(3, 0, f"Дата: {form_data.get('document_date', '')}")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _create_xls_from_invoice_data(self, form_data: Dict) -> BinaryIO:
        """
        Создание XLS файла для Счета на оплату через xlwt
        
        Args:
            form_data: Данные формы Счета
        
        Returns:
            BytesIO с XLS файлом
        """
        # Упрощенная реализация для Счета (можно расширить позже)
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Счет')
        
        style_title = xlwt.XFStyle()
        font_title = xlwt.Font()
        font_title.name = 'Arial'
        font_title.bold = True
        font_title.height = 280  # 14pt
        style_title.font = font_title
        
        ws.write(0, 0, 'СЧЕТ НА ОПЛАТУ', style_title)
        ws.write(2, 0, f"№ {form_data.get('document_number', form_data.get('invoice_number', ''))}")
        ws.write(3, 0, f"Дата: {form_data.get('document_date', form_data.get('invoice_date', ''))}")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _create_xlsx_from_upd_data(self, form_data: Dict) -> BinaryIO:
        """
        Создание XLSX файла для УПД через openpyxl
        
        Этап 2: Полноценный XLSX с точным воспроизведением макета УПД.
        
        Args:
            form_data: Данные формы УПД
        
        Returns:
            BytesIO с XLSX файлом
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "УПД"
        
        # Стили
        header_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        small_font = Font(name='Arial', size=8)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Установка ширины колонок (упрощенная схема, 14 основных колонок)
        column_widths = [4, 30, 8, 8, 12, 12, 12, 15, 12, 12, 8, 8, 8, 15]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        row = 1
        
        # Заголовок документа
        ws.merge_cells(f'A{row}:D{row+2}')
        cell = ws[f'A{row}']
        cell.value = 'Универсальный\nпередаточный\nдокумент'
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        ws.merge_cells(f'E{row}:N{row}')
        cell = ws[f'E{row}']
        doc_number = form_data.get('document_number', '')
        doc_date = form_data.get('document_date', '')
        cell.value = f'Счёт-фактура № {doc_number} от {doc_date}'
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        row += 1
        ws.merge_cells(f'E{row}:N{row}')
        cell = ws[f'E{row}']
        status_map = {'1': '1 — товар (работа, услуга)', '2': '2 — имущественное право'}
        cell.value = f'Статус: {status_map.get(str(form_data.get("status", "1")), "1")}'
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.border = thin_border
        
        row += 2
        
        # Продавец
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].value = 'Продавец:'
        ws[f'A{row}'].font = normal_font
        ws.merge_cells(f'C{row}:N{row}')
        seller = form_data.get('seller', {})
        ws[f'C{row}'].value = f"{seller.get('name', '')} (ИНН: {seller.get('inn', '')}, КПП: {seller.get('kpp', '')})"
        ws[f'C{row}'].font = normal_font
        
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].value = 'Адрес:'
        ws.merge_cells(f'C{row}:N{row}')
        ws[f'C{row}'].value = seller.get('address', '')
        
        row += 2
        
        # Покупатель
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].value = 'Покупатель:'
        ws.merge_cells(f'C{row}:N{row}')
        buyer = form_data.get('buyer', {})
        ws[f'C{row}'].value = f"{buyer.get('name', '')} (ИНН: {buyer.get('inn', '')}, КПП: {buyer.get('kpp', '')})"
        
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].value = 'Адрес:'
        ws.merge_cells(f'C{row}:N{row}')
        ws[f'C{row}'].value = buyer.get('address', '')
        
        row += 2
        
        # Заголовки таблицы товаров
        headers = [
            '№\nп/п',
            'Наименование товара\n(описание работ, услуг)',
            'Код\nвида\nтовара',
            'Единица\nизмерения',
            'Количество\n(объем)',
            'Цена\n(тариф) за\nединицу',
            'Стоимость\nтоваров\nбез НДС',
            'В том числе\nсумма акциза',
            'Налоговая\nставка',
            'Сумма\nНДС',
            'Стоимость\nтоваров\nс НДС',
            'Страна\nпроисхождения',
            'Регистрационный\nномер таможенной\nдекларации',
            'Прослеживаемость'
        ]
        
        header_row = row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(header_row, col_idx, header)
            cell.font = Font(name='Arial', size=8, bold=True)  # Исправлено: добавлен bold
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        row += 1
        
        # Строки товаров
        items = form_data.get('items', [])
        for idx, item in enumerate(items, start=1):
            ws.cell(row, 1, idx).alignment = center_alignment
            ws.cell(row, 1).border = thin_border
            ws.cell(row, 1).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 2, item.get('name', '')).alignment = left_alignment
            ws.cell(row, 2).border = thin_border
            ws.cell(row, 2).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 3, '').alignment = center_alignment
            ws.cell(row, 3).border = thin_border
            ws.cell(row, 3).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 4, item.get('unit_name', 'шт')).alignment = center_alignment
            ws.cell(row, 4).border = thin_border
            ws.cell(row, 4).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 5, float(item.get('quantity', 0))).alignment = right_alignment
            ws.cell(row, 5).number_format = '0.00'
            ws.cell(row, 5).border = thin_border
            ws.cell(row, 5).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 6, float(item.get('price', 0))).alignment = right_alignment
            ws.cell(row, 6).number_format = '#,##0.00'
            ws.cell(row, 6).border = thin_border
            ws.cell(row, 6).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 7, float(item.get('amount_without_vat', 0))).alignment = right_alignment
            ws.cell(row, 7).number_format = '#,##0.00'
            ws.cell(row, 7).border = thin_border
            ws.cell(row, 7).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 8, '—').alignment = center_alignment
            ws.cell(row, 8).border = thin_border
            ws.cell(row, 8).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 9, item.get('vat_rate', 'Без налога')).alignment = center_alignment
            ws.cell(row, 9).border = thin_border
            ws.cell(row, 9).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 10, float(item.get('vat_amount', 0))).alignment = right_alignment
            ws.cell(row, 10).number_format = '#,##0.00'
            ws.cell(row, 10).border = thin_border
            ws.cell(row, 10).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 11, float(item.get('amount_with_vat', 0))).alignment = right_alignment
            ws.cell(row, 11).number_format = '#,##0.00'
            ws.cell(row, 11).border = thin_border
            ws.cell(row, 11).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 12, item.get('country_name', '')).alignment = center_alignment
            ws.cell(row, 12).border = thin_border
            ws.cell(row, 12).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 13, item.get('customs_declaration', '')).alignment = center_alignment
            ws.cell(row, 13).border = thin_border
            ws.cell(row, 13).font = normal_font  # Добавлен шрифт
            
            ws.cell(row, 14, '').alignment = center_alignment
            ws.cell(row, 14).border = thin_border
            ws.cell(row, 14).font = normal_font  # Добавлен шрифт
            
            row += 1
        
        # Итоговая строка
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].value = 'Всего к оплате'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = right_alignment
        ws[f'A{row}'].border = thin_border
        
        ws.cell(row, 7, float(form_data.get('total_amount_without_vat', 0))).alignment = right_alignment
        ws.cell(row, 7).number_format = '#,##0.00'
        ws.cell(row, 7).font = header_font
        ws.cell(row, 7).border = thin_border
        
        ws.cell(row, 8, '—').alignment = center_alignment
        ws.cell(row, 8).border = thin_border
        
        ws.cell(row, 9, 'X').alignment = center_alignment
        ws.cell(row, 9).border = thin_border
        
        ws.cell(row, 10, float(form_data.get('total_vat_amount', 0))).alignment = right_alignment
        ws.cell(row, 10).number_format = '#,##0.00'
        ws.cell(row, 10).font = header_font
        ws.cell(row, 10).border = thin_border
        
        ws.cell(row, 11, float(form_data.get('total_amount_with_vat', 0))).alignment = right_alignment
        ws.cell(row, 11).number_format = '#,##0.00'
        ws.cell(row, 11).font = header_font
        ws.cell(row, 11).border = thin_border
        
        row += 2
        
        # Подписи
        seller_signer = form_data.get('seller_signer', {})
        if seller_signer and seller_signer.get('name'):
            ws[f'A{row}'].value = 'Руководитель организации (продавец):'
            ws[f'A{row}'].font = normal_font
            ws[f'C{row}'].value = f"{seller_signer.get('title', '')} / {seller_signer.get('name', '')}"
            ws[f'C{row}'].font = normal_font
        
        row += 2
        buyer_signer = form_data.get('buyer_signer', {})
        if buyer_signer and buyer_signer.get('name'):
            ws[f'A{row}'].value = 'Руководитель организации (покупатель):'
            ws[f'A{row}'].font = normal_font
            ws[f'C{row}'].value = f"{buyer_signer.get('title', '')} / {buyer_signer.get('name', '')}"
            ws[f'C{row}'].font = normal_font
        
        # Сохранение в BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _create_xlsx_from_akt_data(self, form_data: Dict) -> BinaryIO:
        """
        Создание XLSX файла для Акта выполненных работ
        
        Args:
            form_data: Данные формы Акта
        
        Returns:
            BytesIO с XLSX файлом
        """
        # Упрощенная реализация для Акта
        # TODO: Полная реализация по аналогии с УПД
        wb = Workbook()
        ws = wb.active
        ws.title = "Акт"
        
        ws['A1'] = 'АКТ ВЫПОЛНЕННЫХ РАБОТ (УСЛУГ)'
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        
        ws['A3'] = f"№ {form_data.get('document_number', '')}"
        ws['A4'] = f"Дата: {form_data.get('document_date', '')}"
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _create_xlsx_from_invoice_data(self, form_data: Dict) -> BinaryIO:
        """
        Создание XLSX файла для Счета на оплату
        
        Args:
            form_data: Данные формы Счета
        
        Returns:
            BytesIO с XLSX файлом
        """
        # Упрощенная реализация для Счета
        # TODO: Полная реализация по аналогии с УПД
        wb = Workbook()
        ws = wb.active
        ws.title = "Счет"
        
        ws['A1'] = 'СЧЕТ НА ОПЛАТУ'
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        
        ws['A3'] = f"№ {form_data.get('document_number', form_data.get('invoice_number', ''))}"
        ws['A4'] = f"Дата: {form_data.get('document_date', form_data.get('invoice_date', ''))}"
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _get_excel_filename(self, metadata: Dict, extension: str) -> str:
        """
        Формирование имени файла для Excel
        
        Args:
            metadata: Метаданные документа
            extension: Расширение файла (xls или xlsx)
        
        Returns:
            Имя файла (например: "UPD_125_20260118.xls")
        """
        doc_type = metadata.get('type', 'document').upper()
        doc_number = metadata.get('document_number', '0')
        doc_date = metadata.get('document_date', '').replace('-', '').replace('.', '')
        
        # Очистка имени файла от недопустимых символов
        doc_number = str(doc_number).replace('/', '_').replace('\\', '_')
        
        return f"{doc_type}_{doc_number}_{doc_date}.{extension}"
