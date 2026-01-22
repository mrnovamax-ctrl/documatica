"""
API endpoints для товаров и услуг
"""

import json
import os
import io
from pathlib import Path
from datetime import datetime
import uuid
from typing import List, Optional

import jwt
from fastapi import APIRouter, HTTPException, Header, Cookie, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

router = APIRouter()

# Настройки JWT (должны совпадать с auth.py)
SECRET_KEY = os.getenv("SECRET_KEY", "documatica-secret-key-change-in-production")
ALGORITHM = "HS256"

# Путь к файлам данных
DATA_DIR = Path(__file__).parent.parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)

PRODUCTS_FILE = DATA_DIR / "products.json"


def get_user_id_from_token(authorization: Optional[str] = None, access_token: Optional[str] = None) -> Optional[int]:
    """Извлекает user_id из JWT токена. Возвращает None если токен отсутствует или невалиден."""
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization.split(" ")[1]
    elif access_token:
        token = access_token
    
    if not token:
        return None
    
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id_str = payload.get("sub") or payload.get("user_id")
        return int(user_id_str) if user_id_str else None
    except (jwt.ExpiredSignatureError, jwt.InvalidTokenError, ValueError, TypeError):
        return None


class ProductBase(BaseModel):
    type: str = "product"  # product или service
    name: str
    sku: Optional[str] = None
    unit: str = "шт"
    unit_code: Optional[str] = None
    price: float = 0
    vat_rate: int = 20  # Ставка НДС: 20, 10, 0, -1 (без НДС)
    qty: float = 1
    country: Optional[str] = "Россия"
    country_code: Optional[str] = "643"
    description: Optional[str] = None


class ProductCreate(ProductBase):
    pass


class ProductUpdate(ProductBase):
    name: Optional[str] = None


class Product(ProductBase):
    id: str
    created_at: str
    updated_at: str


def load_products() -> List[dict]:
    """Загрузка товаров из файла"""
    if PRODUCTS_FILE.exists():
        with open(PRODUCTS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def save_products(products: List[dict]):
    """Сохранение товаров в файл"""
    with open(PRODUCTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)


@router.get("/products/")
@router.get("/products", response_model=List[Product])
async def get_products(
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    """Получить список товаров и услуг пользователя"""
    user_id = get_user_id_from_token(authorization, access_token)
    products = load_products()
    
    # Фильтруем по user_id
    if user_id is not None:
        return [p for p in products if p.get("user_id") == user_id]
    return []  # Без авторизации возвращаем пустой список


@router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str):
    """Получить товар по ID"""
    products = load_products()
    for product in products:
        if product["id"] == product_id:
            return product
    raise HTTPException(status_code=404, detail="Товар не найден")


@router.post("/products/")
@router.post("/products", response_model=Product)
async def create_product(
    product: ProductCreate, 
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    """Создать новый товар/услугу"""
    user_id = get_user_id_from_token(authorization, access_token)
    products = load_products()
    
    now = datetime.now().isoformat()
    new_product = {
        "id": str(uuid.uuid4()),
        "type": product.type,
        "name": product.name,
        "sku": product.sku,
        "unit": product.unit,
        "unit_code": product.unit_code,
        "price": product.price,
        "vat_rate": product.vat_rate,
        "qty": product.qty,
        "country": product.country,
        "country_code": product.country_code,
        "description": product.description,
        "user_id": user_id,
        "created_at": now,
        "updated_at": now
    }
    
    products.append(new_product)
    save_products(products)
    
    return new_product


@router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, product: ProductUpdate):
    """Обновить товар/услугу"""
    products = load_products()
    
    for i, existing in enumerate(products):
        if existing["id"] == product_id:
            # Обновляем только переданные поля
            update_data = product.dict(exclude_unset=True)
            for key, value in update_data.items():
                if value is not None:
                    existing[key] = value
            existing["updated_at"] = datetime.now().isoformat()
            
            save_products(products)
            return existing
    
    raise HTTPException(status_code=404, detail="Товар не найден")


@router.delete("/products/{product_id}")
async def delete_product(product_id: str):
    """Удалить товар/услугу"""
    products = load_products()
    
    for i, product in enumerate(products):
        if product["id"] == product_id:
            products.pop(i)
            save_products(products)
            return {"message": "Товар удален"}
    
    raise HTTPException(status_code=404, detail="Товар не найден")


@router.post("/products/bulk", response_model=List[Product])
async def create_products_bulk(
    products_data: List[ProductCreate], 
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    """Массовое создание товаров (для импорта из УПД)"""
    user_id = get_user_id_from_token(authorization, access_token)
    products = load_products()
    new_products = []
    now = datetime.now().isoformat()
    
    for product in products_data:
        # Проверяем, нет ли уже товара с таким же названием для данного пользователя
        existing = next((p for p in products if p["name"].lower() == product.name.lower() and p.get("user_id") == user_id), None)
        
        if not existing:
            new_product = {
                "id": str(uuid.uuid4()),
                "type": product.type,
                "name": product.name,
                "sku": product.sku,
                "unit": product.unit,
                "unit_code": product.unit_code,
                "price": product.price,
                "qty": product.qty,
                "country": product.country,
                "country_code": product.country_code,
                "description": product.description,
                "user_id": user_id,
                "created_at": now,
                "updated_at": now
            }
            products.append(new_product)
            new_products.append(new_product)
    
    if new_products:
        save_products(products)
    
    return new_products


@router.get("/products/search/{query}")
async def search_products(
    query: str, 
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    """Поиск товаров по названию или артикулу"""
    user_id = get_user_id_from_token(authorization, access_token)
    products = load_products()
    query_lower = query.lower()
    
    # Фильтруем по user_id
    if user_id is not None:
        products = [p for p in products if p.get("user_id") == user_id]
    else:
        return []  # Без авторизации возвращаем пустой список
    
    results = [
        p for p in products 
        if query_lower in p["name"].lower() or 
           (p.get("sku") and query_lower in p["sku"].lower())
    ]
    
    return results[:20]  # Максимум 20 результатов


# ============== ИМПОРТ/ЭКСПОРТ XLS ==============

@router.get("/products/template/download")
async def download_products_template():
    """Скачать шаблон XLS для импорта товаров"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары и услуги"
    
    # Заголовки
    headers = ["Наименование", "Артикул", "Ед. изм.", "Цена", "Кол-во", "Страна", "Описание"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Примеры данных
    examples = [
        ["Консультация по налогам", "USL-001", "усл", 5000, 1, "Россия", "Консультация бухгалтера"],
        ["Ведение бухгалтерии", "USL-002", "мес", 15000, 1, "Россия", "Ежемесячное обслуживание"],
        ["Подготовка декларации", "USL-003", "шт", 3000, 1, "Россия", "3-НДФЛ, УСН и др."],
    ]
    
    for row_num, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 35
    
    # Добавляем лист с единицами измерения
    ws2 = wb.create_sheet("Единицы измерения")
    units = [
        ["Код", "Сокращение", "Наименование"],
        ["796", "шт", "Штука"],
        ["876", "усл", "Условная единица (услуга)"],
        ["356", "час", "Час"],
        ["166", "кг", "Килограмм"],
        ["006", "м", "Метр"],
        ["055", "м2", "Квадратный метр"],
        ["113", "м3", "Кубический метр"],
        ["112", "л", "Литр"],
        ["839", "компл", "Комплект"],
        ["421", "мес", "Месяц"],
    ]
    for row_num, row_data in enumerate(units, 1):
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col, value=value)
            if row_num == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = thin_border
    
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 25
    
    # Сохраняем в буфер
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_template.xlsx"}
    )


@router.post("/products/import")
async def import_products_from_xls(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    """Импорт товаров из XLS файла"""
    user_id = get_user_id_from_token(authorization, access_token)
    if not user_id:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Проверяем тип файла
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате .xlsx или .xls")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        products = load_products()
        new_products = []
        now = datetime.now().isoformat()
        
        # Единицы измерения для поиска кода
        unit_codes = {
            "шт": "796", "штука": "796",
            "усл": "876", "услуга": "876",
            "час": "356",
            "кг": "166", "килограмм": "166",
            "м": "006", "метр": "006",
            "м2": "055", "кв.м": "055",
            "м3": "113", "куб.м": "113",
            "л": "112", "литр": "112",
            "компл": "839", "комплект": "839",
            "мес": "421", "месяц": "421",
        }
        
        # Пропускаем заголовок (первую строку)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:  # Пропускаем пустые строки
                continue
            
            name = str(row[0]).strip() if row[0] else ""
            if not name:
                continue
            
            sku = str(row[1]).strip() if len(row) > 1 and row[1] else None
            unit = str(row[2]).strip() if len(row) > 2 and row[2] else "шт"
            
            try:
                price = float(row[3]) if len(row) > 3 and row[3] else 0
            except (ValueError, TypeError):
                price = 0
            
            try:
                qty = float(row[4]) if len(row) > 4 and row[4] else 1
            except (ValueError, TypeError):
                qty = 1
            
            country = str(row[5]).strip() if len(row) > 5 and row[5] else "Россия"
            description = str(row[6]).strip() if len(row) > 6 and row[6] else None
            
            # Ищем код единицы измерения
            unit_lower = unit.lower()
            unit_code = unit_codes.get(unit_lower, "796")  # По умолчанию штуки
            
            # Проверяем, нет ли уже такого товара
            existing = next(
                (p for p in products if p["name"].lower() == name.lower() and p.get("user_id") == user_id), 
                None
            )
            
            if not existing:
                new_product = {
                    "id": str(uuid.uuid4()),
                    "type": "service" if unit_lower in ["усл", "услуга", "час", "мес", "месяц"] else "product",
                    "name": name,
                    "sku": sku,
                    "unit": unit,
                    "unit_code": unit_code,
                    "price": price,
                    "qty": qty,
                    "country": country,
                    "country_code": "643" if country.lower() == "россия" else None,
                    "description": description,
                    "user_id": user_id,
                    "created_at": now,
                    "updated_at": now
                }
                products.append(new_product)
                new_products.append(new_product)
        
        if new_products:
            save_products(products)
        
        return {
            "success": True,
            "imported": len(new_products),
            "message": f"Импортировано {len(new_products)} товаров/услуг"
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка чтения файла: {str(e)}")
